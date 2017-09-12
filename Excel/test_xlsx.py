#!/usr/bin/python
# -*- coding: utf-8 -*-
from openpyxl.styles.named_styles import NamedStyle

__author__ = 'Emmanuel Barillot'

################################################################################
# Petites expériences avec openpyxl
#
# Documentation ici: https://openpyxl.readthedocs.io/en/default/
# doc sur les worksheets ici https://openpyxl.readthedocs.io/en/default/api/openpyxl.worksheet.worksheet.html
#
################################################################################

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side

# creates new workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42
ws['B1'] = 54
ws['C1'] = 'Res'

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()

# acces a cell using row and column notation
d = ws.cell(row=4, column=2, value=10)

# dimensions
print("ws.max_column: {}".format(ws.max_column))
print("dimensions: {}".format(ws.calculate_dimension()))
print("min_column: {}".format(ws.min_column))
print("max_column: {}".format(ws.max_column))
print("min_row: {}".format(ws.min_row))
print("max_row: {}".format(ws.max_row))


# creation d'un style personnel nommé, de façon à l'affecter à une ligne ou une colonne entière
highlight = NamedStyle(name="highlight")
highlight.font = Font(bold=True, size=20)
bd = Side(style='thin', color="000000")
highlight.border = Border(left=None, top=bd, right=None, bottom=bd)
wb.add_named_style(highlight)


for cells in ws.iter_rows(min_row=1,max_row=1,min_col=1,max_col=ws.max_column):
    for cell in cells:
        cell.style = 'highlight'

# row = ws.row_dimensions[1]
# row.style = 'highlight'
# row.style = 'Title'
# row.style = 'Headline 1'
# row.font = Font(underline="single")

# largeur des colonnes pour voir la totalité des valeurs
column_widths = []
for row in ws.iter_rows():
    for i, cell in enumerate(row):
        try:
            column_widths[i] = max(column_widths[i], len(str(cell.value)))
        except IndexError:
            column_widths.append(len(str(cell.value)))

# correction des largeurs de colonnes
column_widths[1] = 10
column_widths[2] = 12

for i, column_width in enumerate(column_widths):
    ws.column_dimensions[get_column_letter(i + 1)].width = column_width

# figer les volets sur la première ligne
freeze_cell = ws.cell(row=2, column=1)
ws.freeze_panes = freeze_cell

# Save the file
wb.save(r"sample.xlsx")

# on relit le fichier
wb2 = load_workbook(r'sample.xlsx')
print(wb2.get_sheet_names())