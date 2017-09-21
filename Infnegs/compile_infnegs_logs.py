# -*- coding: utf-8 -*-

#
# Compilation des compteurs des fichiers logs
#
#

from __future__ import print_function, unicode_literals

import logging
import os
import re

from openpyxl import Workbook as ExcelWorkbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.styles.named_styles import NamedStyle
from openpyxl.utils import get_column_letter
from typing import Iterable, List, Union, Dict
import json

from EBCommons.json_helper import json_navigate
from EBCommons.paths_and_files import filter_files_with_patterns_and_extensions
from EBCommons.prog_helper import LocalError, get_fun_ref, log_exit, log_init, log_write

__author__ = 'Emmanuel Barillot'

# constantes
# pattern pour les noms de fichiers à rechercher (syntaxe analogue au ls du shell Unix)
PATTERN_INGNEGS_LOG = r"chg*.log"
# encoding des fichiers log à analyser, par defaut
DEFAULT_LOG_ENCODING = r'iso-8859-15'


class CompteurOne(object):
    """
    Contient les informations d'un compteur
    """
    __slots__ = ['_num', '_name', '_value']

    def __init__(self, num, name, value):
        # type: (int, unicode, int) -> None
        self._num = num
        self._name = name  # type: unicode
        self._value = value

    def num(self):
        return self._num

    def name(self):
        return self._name

    def value(self):
        return self._value

    def to_tuple(self):
        return self._num, self._name, self._value

    def __unicode__(self):
        # type: () -> unicode
        # comme on travaille en unicode dans ce module, c'est cette fonction qu'il faut surcharger
        return "[{0[0]:>2}] {0[1]:60.60s}: {0[2]:>9}".format(self.to_tuple())

    def __str__(self):
        # type: () -> unicode
        #  la fonction str() fonctionne par defaut en ascii (sys.getdefaultencoding())
        return self.__unicode__()

    def to_str(self):
        # type: () -> unicode
        return self.__unicode__()

    def key(self):
        return self._num


class CompteursFichier(object):
    """
    Contient tous les compteurs associés à un nom de fichier log
    """
    __slots__ = ['_corresp', '_fichier', '_remettant', '_date_run', '_compteurs']

    def __init__(self, fichier, remettant=None, compteurs=None, date_run=None):
        # type: (unicode, unicode, List[CompteurOne]) -> None
        self._fichier = fichier
        self._remettant = remettant
        self._date_run = date_run
        self._compteurs = list(compteurs) if compteurs else list()  # type: List[CompteurOne]

    def set_remettant(self, remettant):
        # type: (unicode) -> None
        self._remettant = remettant

    def remettant(self):
        # type: () -> unicode
        return self._remettant

    def set_date_run(self, date_run):
        # type: (unicode) -> None
        self._date_run = date_run

    def date_run(self):
        # type: () -> unicode
        return self._date_run

    def fichier(self):
        # type: () -> unicode
        return os.path.basename(self._fichier)

    def compteurs(self):
        # type: () -> Iterable[CompteurOne]
        return self._compteurs

    def add(self, compteur_one):
        # type: (CompteurOne) -> List[CompteurOne]
        self._compteurs.append(compteur_one)
        return self._compteurs

    def print_compteurs(self):
        # type: () -> None
        """affichage des compteurs"""
        log_write("Compteurs du fichier {} - {} :".format(self._fichier.encode('utf8'), self._remettant),
                  level=logging.DEBUG)
        for cpt in self._compteurs:
            log_write(cpt.__str__(), level=logging.DEBUG)

    def get_compteurs_as_dict(self):
        # type: () -> dict
        """
        :return: dictionnaire des compteurs du fichier
        """
        # compteurs_dict = dict()
        # for compteur in self._compteurs:
        #     compteur_name = compteur.name()
        #     compteurs_dict[compteur_name] = compteur.value()
        compteurs_dict = dict([(compteur.name(), compteur.value()) for compteur in self._compteurs])
        return compteurs_dict

    @classmethod
    def normalize(cls, compteurs_fichier_list_in):
        # type: (List[CompteursFichier]) -> List[CompteursFichier]
        """
        Normalise les séries de compteurs associés à un ensemble de fichiers
        Tous les fichiers auront artificiellement une valeur (éventuellement 0)
        pour l'ensemble des compteurs présents dans tous les fichiers.
        :param compteurs_fichier_list_in: collection de CompteursFichier
        :return: une collection de CompteursFichier normalisée
        """
        compteur_name_set = set([compteur.name() for compteurs_fichier in compteurs_fichier_list_in
                                 for compteur in compteurs_fichier.compteurs()])
        compteurs_fichier_list_out = list()
        for compteurs_fichier in compteurs_fichier_list_in:
            compteurs_normalises_list = list()
            compteurs_as_dict = compteurs_fichier.get_compteurs_as_dict()
            compteur_num = -1
            for compteur_name in compteur_name_set:
                compteur_num += 1
                compteur_value = compteurs_as_dict[compteur_name] if compteur_name in compteurs_as_dict.keys() else 0
                compteurs_normalises_list += [CompteurOne(compteur_num, compteur_name, compteur_value)]
            compteurs_fichier_list_out += [CompteursFichier(compteurs_fichier.fichier(),
                                                            compteurs_fichier.remettant(),
                                                            compteurs_normalises_list,
                                                            compteurs_fichier.date_run())]
        return compteurs_fichier_list_out

    @classmethod
    def normalize_corresp(cls, compteurs_fichier_list_in, corresp=None):
        # type: (List[CompteursFichier], CompteursCorrespondance) -> List[CompteursFichier]
        """
        Normalise les séries de compteurs associés à un ensemble de fichiers
        Tous les fichiers auront artificiellement une valeur (éventuellement 0)
        pour l'ensemble des compteurs présents dans tous les fichiers.
        :param compteurs_fichier_list_in: collection de CompteursFichier
        :param corresp: correspondances de compteurs
        :return: une collection de CompteursFichier normalisée
        """
        if corresp is None:
            return cls.normalize(compteurs_fichier_list_in)

        compteurs_fichier_list_out = list()
        corresp_compteurs_norm = corresp.get_compteurs_norm()
        for compteurs_fichier in compteurs_fichier_list_in:
            corresp_for_this_fichier = corresp.get_corresp_by_key(compteurs_fichier.fichier())
            compteurs_normalises_dict = dict()
            #  1ere passe pour créer la liste / réceptacle des compteurs normalisés pour un fichier donné
            for compteur_name_norm in corresp_compteurs_norm.keys():
                compteur_num = corresp_compteurs_norm[compteur_name_norm]
                compteurs_normalises_dict[compteur_name_norm] = {"compteur_num"      : compteur_num,
                                                                 "compteur_name_norm": compteur_name_norm,
                                                                 "compteur_value"    : 0}
            compteurs_du_fichier_as_dict = compteurs_fichier.get_compteurs_as_dict()
            # 2eme passe pour attribuer une valeurs aux compteurs normalisés à partir du compteur réel trouvé
            # dans le fichier log
            # on contruit une liste de compteurs pour chaque fichier
            # selon le type de fichier, la correspondance adéquate est utilisée.
            # Chaque nom de compteur trouvé dans le fichier est alors traduit dans son nom normalisé issu de la
            # correspondance
            for compteur_name in compteurs_du_fichier_as_dict.keys():
                if compteur_name in corresp_for_this_fichier.keys():
                    # on met à jour la valeur du compteur normalisé
                    compteur_name_norm = corresp_for_this_fichier[compteur_name]
                    if compteur_name_norm is not None and compteur_name_norm != '':
                        compteurs_normalises_dict[compteur_name_norm]["compteur_value"] = \
                            compteurs_du_fichier_as_dict[compteur_name]

            # on transforme la structure de la liste des compteurs liés à une fichier donné
            compteurs_normalises_list = [CompteurOne(compteurs_normalises_dict[cpt]["compteur_num"],
                                                     compteurs_normalises_dict[cpt]["compteur_name_norm"],
                                                     compteurs_normalises_dict[cpt]["compteur_value"])
                                         for cpt in compteurs_normalises_dict.keys()]
            compteurs_normalises_list.sort(key=CompteurOne.key)
            compteurs_fichier_list_out += [CompteursFichier(compteurs_fichier.fichier(),
                                                            compteurs_fichier.remettant(),
                                                            compteurs_normalises_list,
                                                            compteurs_fichier.date_run())]
        return compteurs_fichier_list_out


class CompteursCorrespondance(object):
    """
    Contient une correspondance entre noms de compteurs. Une correspondance est donc un ensemble d'ensembles de
    compteurs.
    Une clé, c'es à dire un nom de correspondance est associé à chaque ensemble de compteur.
    Un ensemble de compteurs correspond aux compteurs présents dans un fichier log (ou à un sous ensemble de ceux-ci)

    A chaque nom d'un compteur présent dans un fichier log,
    on associe un nouveau nom de compteur qui apparaitra dans le fichier de synthèse. Ca permet de générer une synthèse
    des compteurs tout en gérant l'évolution des noms de compteurs au fil des évolutions des programmes de chargement.
    """
    __slots__ = ['_corresp']

    def __init__(self, json_obj):
        # type: (Dict) -> CompteursCorrespondance
        self._corresp = json_obj

    @classmethod
    def init_from_file(cls, file_name):
        # type: (unicode, unicode) -> CompteursCorrespondance
        """
        Construit une instance d'une correspondance de compteurs à partir d'un fichier de paramétrage
        :param file_name: nom du fichier qui contient la correspondance entre les noms de compteurs
        :return: la correspondance, en fait un dictionnaire de plusieurs correspondances vers un seul ensemble
        de compteurs.
        """
        with open(file_name) as f:
            json_obj = json.load(f)
        return cls(json_obj=json_obj)  # appelle __init__()

    def _get_corresp_key(self, guess_corresp_key):
        # type: (unicode) -> unicode
        """
        Retourne le nom de la correspondance dans le dictionnaire des correspondances.
        la valeur de guess_corresp_key doit contenir le nom de la correspondance.
        Ex: guess_corresp_key = 'chgInfnegs_2017-09-10.log'
            key = 'chgInfnegs'
        :param guess_corresp_key: nom de la correspondance à trouver
        :return: la première clé trouvée dans le dictionnaire des correspondances
        """
        for the_key in self._corresp.keys():
            if the_key in guess_corresp_key:  # on cherche la clé qui ressemble
                return the_key

    def get_corresp_compteur(self, compteur_name, guess_corresp_key):
        # type: (unicode) -> unicode
        """
        Retourne la traduction / correspondance d'un compteur vers un autre nom de compteur normalisé
        :param compteur_name: nom du compteur à traduire
        :param guess_corresp_key: la clé de la correspondance à utiliser
        :return: nom normalisé du compteur
        """
        return self._corresp[self._get_corresp_key(guess_corresp_key)][compteur_name]

    def get_corresp_by_key(self, guess_corresp_key):
        # type: (unicode) -> Dict
        """
        Retourne une correspondance à partir d'un nom qui contient une clé,
         c'est un dictionnaire de noms de compteurs à traduire
        :param guess_corresp_key: la clé de la correspondance à utiliser
        :return: une correspondance
        """
        return self._corresp[self._get_corresp_key(guess_corresp_key)]

    def get_corresp_compteur_names(self, guess_corresp_key):
        # type: (unicode) -> List[unicode]
        """
        Retourne la liste des clés / noms des compteurs à traduire, associées à une correspondance
        :param guess_corresp_key: nom de la correspondance à trouver
        :return: liste de noms de compteurs
        """
        return self._corresp[self._get_corresp_key(guess_corresp_key)].keys()

    def get_compteurs_norm(self):
        # type: () -> Dict
        """
        :return: le dictionnaire des compteurs normalisés et de leur ordre
        """
        return self._corresp["compteurs normalisés"]


# TODO: en faire une méthode de CompteursFichier.set_from_file(filename)
def read_prg_log(the_file_name, encoding=DEFAULT_LOG_ENCODING):
    # type: (unicode, unicode) -> CompteursFichier
    """
    Lit les lignes de compteurs d'un fichier log
    et les retourne dans un tuple de tuple de la forme:
        (fichier,(numero,nom,valeur),(...),...)

    :param the_file_name: nom du fichier à analyser
    :param encoding: encodage du fichier à lire
    :return: un tuple des compteurs trouvés
    """

    def parse_date_run(the_line):
        # type: (unicode) -> unicode or None
        re_date_run = re.compile(r"(.*)\|(.*)\|(.*)\|(.*)\|(.*)")  # RE pour trouver la date
        m_date_run = re_date_run.match(the_line)
        if m_date_run is not None:
            remettant = m_date_run.group(1)
            return remettant
        return None

    def parse_remettant(the_line):
        # type: (unicode) -> unicode or None
        re_remettant = re.compile(r".*LOG\|supportcod .* : *(.*)")  # RE pour trouver ligne du remettant
        m_remettant = re_remettant.match(the_line)
        if m_remettant is not None:
            remettant = m_remettant.group(1)
            return remettant
        return None

    def parse_nb_compteurs(the_line):
        # type: (unicode) -> int or None
        # re_nb_compteurs = re.compile(".*LOG\|\[(.*)\] Compteurs")
        re_nb_compteurs = re.compile(r".*LOG\|\[(.*)\] \w+$")  # RE pour trouver ligne du nb de compteur
        m_nb_compteurs = re_nb_compteurs.match(the_line)
        if m_nb_compteurs is not None:
            nb_compteurs_match = m_nb_compteurs.group(1)
            return int(nb_compteurs_match)
        return None

    def parse_compteur(the_line):
        # type: (unicode) -> CompteurOne or None
        re_compteur = re.compile(r".*LOG\|\[(.*)\] (.*) \.+ : *(-?\d*)")  # RE pour trouver les lignes des compteurs
        m_compteur = re_compteur.match(the_line)
        if m_compteur is not None:
            g1_compteur = m_compteur.group(1)  # numero du compteur
            g2_compteur = m_compteur.group(2)  # nom du compteur
            g3_compteur = m_compteur.group(3)  # valeur du compteur
            return CompteurOne(int(g1_compteur), g2_compteur if g2_compteur is not None else None, int(g3_compteur))
        return None

    compteurs_fichier = CompteursFichier(os.path.basename(the_file_name))
    with open(the_file_name, b'rb') as f:
        # on lit ligne à ligne et on cherche dans chaque ligne
        #   le motif de la ligne qui contient le nb de compteurs
        #   ou le motif d'une ligne qui contient un compteur
        # recherche de la date dans la 1ere ligne
        date_run = parse_date_run(f.readline())
        if date_run:
            log_write(the_file_name + ", date_run: " + date_run)
            compteurs_fichier.set_date_run(date_run)
        for line in f:
            line_decoded = line.decode(encoding)  # decodage de la ligne vers utf-8
            # recherche de la ligne qui contient le remettant
            remettant = parse_remettant(line_decoded)
            if remettant:
                log_write(the_file_name + ", remettant: " + remettant)
                compteurs_fichier.set_remettant(remettant)
            # la ligne du nb de compteurs
            nb_compteurs = parse_nb_compteurs(line_decoded)
            if nb_compteurs:
                log_write(the_file_name + ", nb compteurs trouvés: " + str(nb_compteurs))
            # les compteurs
            compteur = parse_compteur(line_decoded)
            if compteur is not None:
                compteurs_fichier.add(compteur)
    return compteurs_fichier


def call_read_prg_log(dir_name, file_name):
    # type: (unicode, unicode) -> bool
    """
    Appel unitaire de la fonction de recherche des compteurs read_prg_log()

    :param dir_name: répertoire du fichier à analyser
    :param file_name: nom du fichier log à analyser
    :return: True si exécution OK
    """
    the_file_name = os.path.join(dir_name, file_name)
    try:
        cpts = read_prg_log(the_file_name)
        cpts.print_compteurs()
        return True
    except Exception as e:
        log_write(LocalError(e).__str__(), level=logging.ERROR)
        return False


def read_prg_log_many(path_src_file, file_name_list, encoding=DEFAULT_LOG_ENCODING):
    # type: (unicode, List[unicode], unicode) -> List[CompteursFichier]
    """
    Lit les lignes de compteurs d'une liste de fichiers log et les retourne dans une liste de CompteursFichier

    :param path_src_file: chemin vers les fichiers à analyser
    :param file_name_list: liste des noms des fichiers à analyser (tous dans le même répertoire path_src)
    :param encoding: encodage à utiliser pour lire les fichiers
    """
    return map(lambda x: read_prg_log(x, encoding),
               [os.path.join(path_src_file, file_name) for file_name in file_name_list])


def call_read_prg_log_many(dir_name, re_file_name):
    # type: (unicode) -> bool
    """
    Appel unitaire de la fonction de recherche des compteurs read_prg_log_many()

    :param dir_name: répertoire à scruter pour trouver les fichiers log à analyser
    :param re_file_name: expression régulière pour les noms de fichiers à analyser
    :return: bool
    """
    try:
        file_name_list = \
            filter_files_with_patterns_and_extensions(os.listdir(dir_name), include_patterns=(re_file_name,))
        cpt_list = read_prg_log_many(dir_name, file_name_list, DEFAULT_LOG_ENCODING)
        for cpt in cpt_list:
            cpt.print_compteurs()
        return True
    except Exception as e:
        log_write(LocalError(e).__str__(), level=logging.ERROR)
        return False


def excel_write_log_cpt(excel_file_full_name, corresp_file_name=None, compteurs_fichier=None,
                        compteurs_fichiers_list=None):
    # type: (unicode, unicode or None, CompteursFichier, Iterable[CompteursFichier]) -> None
    """
    Crée un fichier Excel à partir des compteurs d'un ou de plusieurs fichiers.

    :param excel_file_full_name: nom du fichier Excel en sortie, chemin compris
    :param corresp_file_name: nom du fichier qui contient les correspondances
    :param compteurs_fichier: les compteurs d'un fichier
    :param compteurs_fichiers_list: une liste de compteurs de fichiers
    :return: None
    """

    def add_title_style(wb):
        # creation d'un style personnel nommé, de façon à l'affecter à une ligne ou une colonne entière
        title_style = NamedStyle(name="title_style")
        title_style.font = Font(bold=True, size=14)
        bd = Side(style='thin', color="000000")
        title_style.border = Border(left=None, top=bd, right=None, bottom=bd)
        title_style.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
        wb.add_named_style(title_style)

    # création
    workbook = ExcelWorkbook()
    add_title_style(workbook)
    # grab the active worksheet
    feuil1 = workbook.active
    feuil1.title = "Compteurs"
    # création d'une nouvelle feuille 1
    # feuil1 = workbook.create_sheet('Compteurs')

    if compteurs_fichiers_list is None:
        if compteurs_fichier is None:
            raise Exception('Give either a row or a collection')
        else:
            compteurs_fichiers_list = [compteurs_fichier]

    if corresp_file_name is None:
        compteurs_list = CompteursFichier.normalize(compteurs_fichiers_list)  # type: List[CompteursFichier]
    else:
        compteurs_list = CompteursFichier.normalize_corresp(compteurs_fichiers_list,
                                                            CompteursCorrespondance.init_from_file(
                                                                corresp_file_name))  # type: List[CompteursFichier]

    # ajout des en-têtes, à partir des libellés de la premiere ligne
    # ça suffit car les compteurs ont été normalisés
    title_line = ['Fichier', 'Remettant', 'Date chargement'] + [compteur.name() for compteur in compteurs_list[0].compteurs()]
    feuil1.append(title_line)

    # ajout des valeurs dans les lignes suivantes
    for compteurs_fichier in compteurs_list:
        next_line = [compteurs_fichier.fichier(),
                     compteurs_fichier.remettant(),
                     compteurs_fichier.date_run()] + \
                    [str(compteur.value()) for compteur in compteurs_fichier.compteurs()]
        feuil1.append(next_line)

    # ajustement éventuel de la largeur de chaque colonne
    for i in range(feuil1.max_column):
        feuil1.column_dimensions[get_column_letter(i + 1)].width = 15
    feuil1.column_dimensions[get_column_letter(1)].width = 40

    # figer les volets sur la première ligne
    feuil1.freeze_panes = feuil1.cell(row=2, column=1)

    # application du style au titre
    for cells in feuil1.iter_rows(min_row=1, max_row=1, min_col=1, max_col=feuil1.max_column):
        for cell in cells:
            cell.style = 'title_style'

    # création matérielle du fichier résultant
    try:
        workbook.save(excel_file_full_name)
    except Exception as e:
        log_write(LocalError(e).__str__(), level=logging.ERROR)


def call_excel_write_log_cpt(dir_name, excel_file_name):
    # type: (unicode, unicode) -> bool
    try:
        file_name = os.path.join(dir_name, excel_file_name)
        # compteurs = CompteursFichier(
        #     'fichier_1'
        #     , [CompteurOne(0, 'cpt1', 12)
        #         , CompteurOne(1, 'cpt2', 11)
        #         , CompteurOne(2, 'cpt3', 9)
        #         , CompteurOne(3, 'cpt4', 8)]
        #     )
        compteurs_coll = [CompteursFichier('fichier_1'
                                           , 'Remettant 1'
                                           , map(lambda x: CompteurOne(*x)
                                                 , [
                                                     (' 0', 'cpt1', 12)
                                                     , (' 1', 'cpt2', 11)
                                                     , (' 2', 'cpt3', 9)
                                                     , (' 3', 'cpt4', 8)
                                                     ]
                                                 )
                                           )
            , CompteursFichier('fichier_2'
                               , 'Remettant 2'
                               , map(lambda x: CompteurOne(*x)
                                     , [
                                         (' 0', 'cpt1', 13)
                                         , (' 1', 'cpt2', 15)
                                         , (' 2', 'cpt3', 6)
                                         , (' 3', 'cpt4', 7)
                                         ]
                                     )
                               )
                          ]

        # excel_write_log_cpt(file_name, compteurs_row=compteurs)
        excel_write_log_cpt(file_name, corresp_file_name=None, compteurs_fichiers_list=compteurs_coll)
        return True
    except Exception as e:
        log_write(LocalError(e).__str__(), level=logging.ERROR)
        return False


def call_read_prg_log_to_excel(dir_name, file_name, excel_file_name):
    # type: (unicode, unicode, unicode) -> bool
    """
    Lit un fichier de logs et ecrit les compteurs extraits dans un fichier Excel
    :param dir_name: répertoire où se trouve le fichier
    :param file_name: le nom du fichier
    :param excel_file_name: le nom du fichier Excel
    :return: bool
    """
    the_file_name = os.path.join(dir_name, file_name)
    the_excel_file_name = os.path.join(dir_name, excel_file_name)
    try:
        cpts = read_prg_log(the_file_name)
        cpts.print_compteurs()
        excel_write_log_cpt(the_excel_file_name, compteurs_fichier=cpts)
        return True
    except Exception as e:
        log_write(LocalError(e).__str__(), level=logging.ERROR)
        return False


def call_read_prg_log_to_excel_many(path_src, file_name_re, path_dest, excel_file_name,
                                    corresp_file_name, encoding_src=DEFAULT_LOG_ENCODING):
    # type: (unicode, unicode, unicode, unicode, unicode) -> bool
    """
    Lit une liste de fichiers de logs et ecrit les compteurs extraits dans un fichier Excel
    :param path_src: répertoire où se trouvent les fichiers
    :param file_name_re: le motif des noms de fichiers en expression régulière nom du fichier
    :param encoding_src: encodage des fichiers logs (par ex iso-8859-15 sous Linux)
    :param path_dest: répertoire de destination du fichier Excel
    :param excel_file_name: le nom du fichier Excel produit
    :param corresp_file_name: nom du fichier des correspondances de compteurs
    :return: True si succès
    """
    file_name_re = PATTERN_INGNEGS_LOG if file_name_re is None else file_name_re
    # on recupere les noms de fichiers à traiter
    file_name_list = filter_files_with_patterns_and_extensions(os.listdir(path_src), include_patterns=(file_name_re,))
    try:
        cpt_list = read_prg_log_many(path_src, file_name_list, encoding_src)
        # affichage des compteurs
        if cpt_list:
            for cpt in cpt_list:
                cpt.print_compteurs()
            excel_write_log_cpt(os.path.join(path_dest, excel_file_name), corresp_file_name,
                                compteurs_fichiers_list=cpt_list)
        return True
    except Exception as e:
        log_write(LocalError(e).__str__(), level=logging.ERROR)
        raise
        # return False


#####################
# programme principal
#####################
def launch(path_src, log_file_name, encoding_src, path_dest, excel_file_name, corresp_file_name):
    # type: (unicode, unicode, unicode, unicode, unicode) -> List(bool)
    """
    Lance les fonctions pour test / validation

    :param path_src: repertoire source où se trouvent les fichiers log à analyser
    :param log_file_name: nom du fichier à analysez quand execution sur un seul fichier
    :param encoding_src: encodage des fichiers log
    :param path_dest: répertoire de destination pour le fichier Excel
    :param excel_file_name: nom du fichier Excel
    :param corresp_file_name: nom du fichier des correspondances de compteurs
    :return: liste de bool
    """
    # liste de fonctions à tester
    dict_of_funs_to_test = {
        r'call_read_prg_log'                : (path_src, log_file_name)
        , r'call_excel_write_log_cpt'       : (path_src, excel_file_name)
        , r'call_read_prg_log_to_excel'     : (path_src, log_file_name, path_dest, excel_file_name)
        , r'call_read_prg_log_many'         : (path_src, PATTERN_INGNEGS_LOG)
        , r'call_read_prg_log_to_excel_many': (
            path_src, PATTERN_INGNEGS_LOG, path_dest, excel_file_name, corresp_file_name, encoding_src)
        }

    # liste des références des fonctions
    funs_available = map(lambda x: get_fun_ref(x, __name__), dict_of_funs_to_test.keys())
    # Autre technique qui utilise eval, peut être dangereuse si le nom de la fonction provient d'une entrée utilisateur
    # funs_available = map(eval, dict_of_funs_to_test.keys())

    # la liste des arguments des fonctions testées
    for fun in funs_available:
        if "inspect" not in globals().keys():
            from inspect import getargspec
            args, varargs, varkw, defaults = getargspec(fun)
            log_write("{}:".format(fun))
            log_write("{:10s}: {}".format("args", args))
            log_write("{:10s}: {}".format("varargs", varargs))
            log_write("{:10s}: {}".format("varkw", varkw))
            log_write("{:10s}: {}".format("defaults", defaults))

    # funs_to_test = [call_read_prg_log] # la fonction à tester
    # funs_to_test = [call_excel_write_log_cpt] # la fonction à tester
    # funs_to_test = [call_read_prg_log_to_excel] # la fonction à tester
    # funs_to_test = [call_read_prg_log_many]
    funs_to_test = [call_read_prg_log_to_excel_many]

    # lancement des tests
    call_return = []
    for fun_to_test in funs_to_test:
        log_write(">>>>> testing: {}".format(fun_to_test.func_name))
        # l'étoile devant *dict_of_funs_to_test[fun_to_test.func_name]
        # sert à "unpacker" le tuple stocké dans chaque entrée de ce dictionnaire
        # il faut unpacker le tuple avant d'en passer les éléments à la fonction appelée
        call_return.append((fun_to_test.func_name,
                            "OK" if fun_to_test(*dict_of_funs_to_test[fun_to_test.func_name]) else "KO"))
    return call_return


##################################
# lancement du programme principal
##################################
if __name__ == "__main__":
    log_init(level=logging.DEBUG)
    log_write(">>>>>>>>>>>>>>>>>>>> Nouvelle série de tests <<<<<<<<<<<<<<<<<<<<")

    path_root = r"D:\Documents\Projets\work\Infnegs_logs\2017-08\qqn"
    # path_root = r"D:\Documents\Projets\work\Infnegs_logs\2017-08"
    path_src = path_root
    path_dest_for_excel = path_root
    one_log_file_name = "chgInfnegs_201610031669949.log"
    excel_out_filename = "compteurs.xlsx"
    corresp_file_name = r'compteurs_correspondances.json'

    res = launch(path_src, one_log_file_name, DEFAULT_LOG_ENCODING, path_dest_for_excel, excel_out_filename,
                 corresp_file_name)
    if res is not None:
        for elem in res:
            log_write("{}() => {}".format(unicode(elem[0]), unicode(elem[1])))

    log_exit()
