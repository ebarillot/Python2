# -*- coding: utf-8 -*-
from __future__ import unicode_literals
import inspect
import logging
import os
import pprint
import re
import sys
import traceback
import typing

import paramiko as pk
from openpyxl import Workbook as ExcelWorkbook
from openpyxl.styles import Font, Border, Side
from openpyxl.styles.named_styles import NamedStyle
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

__author__ = 'Emmanuel Barillot'

#### -*- coding: cp1252 -*-
#### -*- coding: iso-8859-15 -*-



def call_stack():
    """
    Produit un résumé de la "stack call"

    :return: Une chaine de caractères qui contient un résumé de la "stack call"
    """
    stack = [frame[3] for frame in inspect.stack()
             if frame[3] not in [inspect.stack()[0][3],"<module>"]]
    s='()->'.join(reversed(stack))
    s += '\n' + traceback.format_exc()
    return s


def connect_to_host(host,username,password):
    try:
        ssh = pk.SSHClient()
        ssh.load_system_host_keys()
        # la commande précédente va lire le fichier C:\Users\emmanuel_barillot\.ssh\known_hosts
        ssh.set_missing_host_key_policy(pk.AutoAddPolicy())
        ssh.connect(host, username=username, password=password)
        logging.info("Connected to %s" % host)
        return ssh
    except pk.AuthenticationException:
        logging.critical("Authentication failed when connecting to %s" % host)
        return None
    except:
        logging.critical("Could not SSH to %s, waiting for it to start" % host)
        return None


def exec_command(ssh, command):
    data = []
    # Send the command (non-blocking)
    stdin, stdout, stderr = ssh.exec_command(command)
    stdin.close()
    for line in stdout.read().splitlines():
        data.append(line)
        # logging.debug('line: %s' % line)
    return data
    # autre méthode qui n'arrive pas à tout récupérer
    # # Wait for the command to terminate
    # while not stdout.channel.exit_status_ready():
    #     # Only print data if there is data to read in the channel
    #     if stdout.channel.recv_ready():
    #         rl, wl, xl = select.select([stdout.channel], [], [], 0.0)
    #         if len(rl) > 0:
    #             # Print data from stdout
    #             data += stdout.channel.recv(1024)
    #             # logging.info(stdout.channel.recv(1024),)


def exists_dest_dir(dest_dir):
    return os.access(dest_dir, os.F_OK)


def access_dest_dir(dest_dir):
    return os.access(dest_dir, os.R_OK|os.W_OK|os.X_OK)


def become_accessible_dest_dir(dest_dir):
    # type: (unicode) -> bool
    access_mode = 0o755
    if os.access(dest_dir, os.F_OK):    # existe
        if os.access(dest_dir, os.R_OK | os.W_OK | os.X_OK):  # existe et accessible
            return True
        else: # existe et non accessible => on essaie de le rendre accessible
            os.chmod(dest_dir, access_mode)
    else: # inexistant => on essaie de le créer
        os.makedirs(dest_dir, access_mode)
    # on vérifie s'il est accesible maintenant
    if os.access(dest_dir, os.R_OK | os.W_OK | os.X_OK): # existe et accessible
        return True
    else:
        return False


def is_accessible_src_dir(ssh,src_dir):
    sftp = ssh.open_sftp()
    sftp_attr = sftp.stat(src_dir)
    if sftp_attr is not None:
        return sftp_attr.st_size > 0
    else:
        return False


def read_host_src_dir_file_names(ssh, src_dir, how):
    if how == "ls":
        cmd = "ls " + src_dir + "/chgInfnegs*.log"
        logging.info(cmd)
        file_names = exec_command(ssh, cmd)
    else:
        sftp = ssh.open_sftp()
        file_names = sftp.listdir(path=".")
    return file_names


def dir_file_names(dir, file_re=".*"):
    """
    Retourne la liste des fichiers réguliers présents dans le répertoire
    :param dir:
    :return:
    """
    file_re_comp = re.compile(file_re)
    # all_file_names = os.listdir(dir)
    # filtered_file_names = []
    # for file_name in all_file_names:
    #     match_file_name = file_re_comp.match(file_name)
    #     if match_file_name is not None:
    #         filtered_file_names.append(file_name)
    filtered_file_names = [file_name for file_name in os.listdir(dir) if file_re_comp.match(file_name) is not None]
    return filtered_file_names


def copy_files_src_to_dest(dest_dir, src_dir, host, host_user, host_password):
    def print_file_names(file_names):
        logging.info("""Find: %d files """ % len(file_names))
        for file_name in file_names:
            logging.info("""%s""" % file_name)

    logging.info("Copy files from {}:{} to {}".format(host, src_dir, dest_dir))
    try:
        if become_accessible_dest_dir(dest_dir):
            logging.info("Destination directory {} is accessible".format(dest_dir))
        else:
            raise Exception("Destination directory {} is not accessible, STOP".format(dest_dir))
    except:
        raise
    try:
        ssh = connect_to_host(host, host_user, host_password)
        if not ssh:
            raise Exception("Connexion refused to remote server " + host)
    except:
        raise
    try:
        if is_accessible_src_dir(ssh, src_dir):
            logging.info("Source directory {} is accessible".format(src_dir))
        else:
            raise Exception("source directory {} is not accessible, STOP".format(dest_dir))
    except:
        raise
    file_names = read_host_src_dir_file_names(ssh, src_dir, "ls")
    file_basenames = [f.split('/')[-1:][0] for f in file_names]
    # print_file_names(file_names2)
    os.chdir(dest_dir)
    sftp = ssh.open_sftp()
    for i in range(len(file_names)):
        already_exists = os.access(file_basenames[i], os.F_OK)
        if already_exists:
            logging.warning("{} already exists".format(file_basenames[i]))
        else:
            logging.info("Copy file {}".format(file_basenames[i]))
            try:
                sftp.get(file_names[i], file_basenames[i])  # remote => local
            except:
                logging.error("Pb copy file {}".format(file_basenames[i]))
    logging.info("Copy done, closing SSH connection")
    ssh.close()
    return True


def call_copy_files_src_to_dest(dest_dir, src_dir, host, host_user, host_password):
    try:
        copy_files_src_to_dest(dest_dir, src_dir, host, host_user, host_password)
        return True
    except Exception as e:
        logging.error(e)
        logging.error(call_stack())
        raise
        # return False


def call_ssh(host, host_user, host_password):
    def print_remote_uname(ssh):
        lines = exec_command(ssh, "uname -a")
        for line in lines:
            logging.info("""%s""" % line)

    def print_remote_HOME(ssh):
        lines = exec_command(ssh, "ls -ltr $HOME")
        for line in lines:
            logging.info("""%s""" % line)

    try:
        ssh = connect_to_host(host, host_user, host_password)
        if not ssh:
            logging.error("Connexion refused to remote server " + host)
            return False
    except Exception as e:
        logging.error(e)
        logging.error(call_stack())
        raise
        # return False
    print_remote_uname(ssh)
    print_remote_HOME(ssh)
    logging.info("Command done, closing SSH connection")
    ssh.close()
    return True


def read_prg_log(the_file_name, encoding=r'iso-8859-15'):
    """
    Lit les lignes de compteurs d'un fichier log et les retourne dans un tuple de tuple
    de la forme:
    (fichier,(numero,nom,valeur),(...),...)
    """
    # re_nb_compteurs = re.compile(".*LOG\|\[(.*)\] Compteurs")
    re_nb_compteurs = re.compile(".*LOG\|\[(.*)\] \w+$")   # RE pour trouver ligne du nb de compteur
    re_compteur = re.compile(".*LOG\|\[(.*)\] (.*) \.+ : *(\d*)")   # RE pour trouver les lignes des compteurs
    cpt = [the_file_name,]
    with open(the_file_name, 'rb') as f:
        for line in f:
            line_decoded = line.decode(encoding)
            # la ligne du nb de compteurs
            m_nb_compteurs = re_nb_compteurs.match(line_decoded)
            g1_nb_compteurs = m_nb_compteurs.group(1) if m_nb_compteurs is not None else None
            # les compteurs
            m_compteur = re_compteur.match(line_decoded)
            if m_compteur is not None:
                g1_compteur = m_compteur.group(1)
                g2_compteur = m_compteur.group(2)
                g3_compteur = m_compteur.group(3)
                cpt.append((g1_compteur, g2_compteur if g2_compteur is not None else None,g3_compteur))
    return tuple(cpt)


def call_read_prg_log(dir_name, file_name):
    the_file_name = dir_name + "\\" + file_name
    try:
        cpt = read_prg_log(the_file_name)
        logging.info("Compteurs du fichier {}:".format(cpt[0].encode('utf8')))
        for i in range(1,len(cpt)):
            logging.info("[{0[0]}] {0[1]:60.60s}: {0[2]:>9}".format(tuple([x.encode('utf8') for x in cpt[i]])))
        return True
    except Exception as e:
        logging.error(e)
        logging.error(call_stack())
        raise
        # return False


def read_prg_log_many(file_name_list, encoding=r'iso-8859-15'):
    """
    Lit les lignes de compteurs d'une liste de fichiers log et les retourne dans une liste de tuple de tuple
    de la forme:
    [ (fichier_1,(numero_cpt1,nom_cpt1,valeur_cpt1)
                ,(numero_cpt2,nom_cpt2,valeur_cpt2)
                ,...)
    , (fichier_2,(numero_cpt1,nom_cpt1,valeur_cpt1)
                ,(numero_cpt2,nom_cpt2,valeur_cpt2)
                ,...)
    , (...)
    ]
    """
    cpt_list = []
    for one_file_name in file_name_list:
        cpt_list.append(read_prg_log(one_file_name, encoding))
    return tuple(cpt_list)


def call_read_prg_log_many(dir_name):
    file_name_list = dir_file_names(dir_name,"^chg.*\.log")
    try:
        cpt_list = read_prg_log_many([dir_name+'/'+file_name for file_name in file_name_list])
        for cpt in cpt_list:
            logging.info("Compteurs du fichier {}:".format(cpt[0]))
            for i in range(1,len(cpt)):
                logging.info("[{0[0]}] {0[1]:60.60s}: {0[2]:>9}".format(cpt[i]))
        return True
    except Exception as e:
        logging.error(e)
        logging.error(call_stack())
        raise
        # return False


def excel_write_log_cpt(excel_file_name, compteurs_row=None, compteurs_rows=None):
    def add_title_style(wb):
        # creation d'un style personnel nommé, de façon à l'affecter à une ligne ou une colonne entière
        title_style = NamedStyle(name="title_style")
        title_style.font = Font(bold=True, size=14)
        bd = Side(style='thin', color="000000")
        title_style.border = Border(left=None, top=bd, right=None, bottom=bd)
        title_style.alignment = Alignment(wrapText=True,horizontal='center',vertical='center')
        wb.add_named_style(title_style)

    # création
    workbook = ExcelWorkbook()
    add_title_style(workbook)
    # grab the active worksheet
    feuil1 = workbook.active
    feuil1.title = "Compteurs"
    # création d'une nouvelle feuille 1
    # feuil1 = workbook.create_sheet('Compteurs')

    if compteurs_row is None and compteurs_rows is None:
        raise Exception('Give either a row or a collection')
    if compteurs_row is not None and compteurs_rows is not None:
        raise Exception('Give either a row or a collection, not both')
    if compteurs_row is not None:
        crows = []
        crows.append(compteurs_row)
    else:
        crows = compteurs_rows

    # ajout des en-têtes, à partir des libellés de la premiere ligne
    crow = crows[0]
    title_line = ['fichier'] + [crow[i][1]  for i in range(1,len(crow))]
    feuil1.append(title_line)

    for i in range(len(crows)):
        # ajout des valeurs dans les lignes suivantes
        crow = crows[i]
        next_line = [crow[0].split('/')[-1]] + [int(crow[i][2])  for i in range(1,len(crow))]
        feuil1.append(next_line)

    # ajustement éventuel de la largeur de chaque colonne
    for i in range(feuil1.max_column):
        feuil1.column_dimensions[get_column_letter(i + 1)].width = 15
    feuil1.column_dimensions[get_column_letter(1)].width = 40

    # figer les volets sur la première ligne
    feuil1.freeze_panes = feuil1.cell(row=2, column=1)

    # application du style au titre
    for cells in feuil1.iter_rows(min_row=1, max_row=1, min_col=1, max_col=feuil1.max_column):
        for cell in cells:
            cell.style = 'title_style'

    # création matérielle du fichier résultant
    workbook.save(excel_file_name)


class Compteurs(object):
    # TODO en construction
    def __init__(self,*compteurs):
        self.cpt = (compteurs)
# TODO ajouter support de l'indexing __getitem__ ?

class Compteurs_row(object):
    # TODO en construction
    def __init__(self,fichier,*compteurs):
        self.row = (fichier,Compteurs(compteurs))


def call_excel_write_log_cpt(dir_name, excel_file_name):
    try:
        file_name = dir_name + "\\" + excel_file_name   # sous windows
        compteurs = (
            'fichier_1'
            , (' 0', 'cpt1', 12)
            , (' 1', 'cpt2', 11)
            , (' 2', 'cpt3', 9)
            , (' 3', 'cpt4', 8)
        )
        compteurs_coll = [
            ('fichier_1'
             , (' 0', 'cpt1', 12)
             , (' 1', 'cpt2', 11)
             , (' 2', 'cpt3', 9)
             , (' 3', 'cpt4', 8)
             )
            , ('fichier_2'
               , (' 0', 'cpt1', 13)
               , (' 1', 'cpt2', 15)
               , (' 2', 'cpt3', 6)
               , (' 3', 'cpt4', 7)
               )
        ]
        # excel_write_log_cpt(file_name, compteurs_row=compteurs)
        excel_write_log_cpt(file_name, compteurs_rows=compteurs_coll)
        return True
    except Exception as e:
        logging.error(e)
        logging.error(call_stack())
        raise
        # return False


def call_read_prg_log_to_excel(dir_name, file_name, excel_file_name):
    """
    Lit un fichier de logs et ecrit les compteurs extraits dans un fichier Excel
    :param dir_name: répertoire où se trouve le fichier
    :param file_name: le nom du fichier
    :return: rien
    """
    the_file_name = dir_name + "\\" + file_name
    the_excel_file_name = dir_name + "\\" + excel_file_name
    try:
        cpt = read_prg_log(the_file_name)
        logging.info("Compteurs du fichier {}:".format(cpt[0]))
        for i in range(1,len(cpt)):
            logging.info("[{0[0]}] {0[1]:60.60s}: {0[2]:>9}".format(cpt[i]))
        excel_write_log_cpt(the_excel_file_name, compteurs_row=cpt)
        return True
    except Exception as e:
        logging.error(e)
        logging.error(call_stack())
        raise
        # return False


def call_read_prg_log_many_to_excel(dir_name, file_name_re, excel_file_name):
    """
    Lit une liste de fichiers de logs et ecrit les compteurs extraits dans un fichier Excel
    :param dir_name: répertoire où se trouvent les fichiers
    :param file_name_re: le motif des noms de fichiers en expression régulière nom du fichier
    :param excel_file_name: le nom du fichier Excel produit
    :return: True si succès
    """
    file_name_re = ".*\.log" if file_name_re is None else file_name_re
    file_name_list = dir_file_names(dir_name,file_re=file_name_re)
    the_excel_file_name = dir_name + "\\" + excel_file_name
    try:
        cpt_list = read_prg_log_many([dir_name + '/' + file_name for file_name in file_name_list])
        for cpt in cpt_list:
            logging.info("Compteurs du fichier {}:".format(cpt[0].encode('utf8')))
            for i in range(1,len(cpt)):
                logging.info("[{0[0]}] {0[1]:60.60s}: {0[2]:>9}".format(cpt[i]))
        excel_write_log_cpt(the_excel_file_name, compteurs_rows=cpt_list)
        return True
    except Exception as e:
        logging.error(e)
        logging.error(call_stack())
        raise
        # return False


#####################
# programme principal
#####################
def main(argv):
    host            = r'frtrsifd01'
    host_user       = r'bfrance'
    host_password   = r'bfrance'
    src_dir         = r"/PROSIP_LOGS/BFRANCE"
    work_dir        = b"C:\Users\emmanuel_barillot\Documents\Work"
    dest_dir        = work_dir + r"\Infnegs_logs\2017-11"
    file_name       = "chgInfnegs_201610311788352.log"
    excel_file_name = "compteurs.xlsx"

    # la liste des arguments des fonctions testées
    funs_available = [
        call_ssh
        , call_copy_files_src_to_dest
        , call_read_prg_log
        , call_excel_write_log_cpt
        , call_read_prg_log_to_excel
        , call_read_prg_log_many
        , call_read_prg_log_many_to_excel
    ]
    for fun in funs_available:
        args, varargs, varkw, defaults = inspect.getargspec(fun)
        logging.info("{}:".format(fun))
        logging.info("{:10s}: {}".format("args", args))
        logging.info("{:10s}: {}".format("varargs", varargs))
        logging.info("{:10s}: {}".format("varkw", varkw))
        logging.info("{:10s}: {}".format("defaults", defaults))

    # de chaque fonction présente dans une liste de fonctions à tester
    dict_of_funs_to_test = {
        'call_ssh': (host, host_user, host_password)
        , 'call_copy_files_src_to_dest': (dest_dir, src_dir, host, host_user, host_password)
        , 'call_read_prg_log': (dest_dir, file_name)
        , 'call_excel_write_log_cpt': (dest_dir, excel_file_name)
        , 'call_read_prg_log_to_excel': (dest_dir, file_name, excel_file_name)
        , 'call_read_prg_log_many': (dest_dir,)
        , 'call_read_prg_log_to_excel_many': (dest_dir, '.*\.log', excel_file_name)
    }

    # TODO vérifier que tous les arguments indispensables sont bien présents
    # en inspectant la liste des arguments de chaque fonction testée

    # funs_to_test = [call_ssh, call_copy_files_src_to_dest,call_read_prg_log]
    # funs_to_test = [call_ssh] # la fonction
    funs_to_test = [call_copy_files_src_to_dest] # la fonction
    # funs_to_test = [call_read_prg_log] # la fonction à tester
    # funs_to_test = [call_excel_write_log_cpt] # la fonction à tester
    # funs_to_test = [call_read_prg_log_to_excel] # la fonction à tester
    # funs_to_test = [call_read_prg_log_many]
    # funs_to_test = [call_read_prg_log_many_to_excel]

    # lancement des tests
    call_return = []
    for fun_to_test in funs_to_test:
        logging.info(">>>>> testing: {}".format(fun_to_test.func_name))
        call_return.append((fun_to_test.func_name,
                            "OK" if fun_to_test(*dict_of_funs_to_test[fun_to_test.func_name]) else "KO"))
    return call_return


def main_logging():
    # logger = logging.getLogger('root')
    LOG_FILE   = 'call_ssh_utf8.log'
    LOG_FORMAT = "%(asctime)s|%(levelname)-7s|%(name)s|%(filename)s|%(funcName)-20s - %(lineno)s|%(message)s"
    logging.basicConfig(filename=LOG_FILE, format=LOG_FORMAT, level=logging.INFO)

# TODO ecrire une fonction qui prend une str en entrée, teste son encodage (si elle est unicode) et qui la convertit
# dans l'encodage du fichier log

##################################
# lancement du programme principal
##################################
if __name__ == "__main__":
    main_logging()
    logging.info(">>>>>>>>>>>>>>>>>>>> Nouvelle série de tests <<<<<<<<<<<<<<<<<<<<")
    pp = pprint.PrettyPrinter(indent=4)
    pp.pprint("Local path: {}".format(os.environ['HOMEPATH']))
    res = main(sys.argv)
    if res is not None:
        for elem in res:
            logging.info("{}() => {}".format(str(elem[0]),str(elem[1])))

