# -*- coding: utf-8 -*-

#
# Compilation des compteurs des fichiers logs
#
#

from __future__ import print_function, unicode_literals

import logging
import os
import re
from collections import OrderedDict

from openpyxl import Workbook as ExcelWorkbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.styles.named_styles import NamedStyle
from openpyxl.utils import get_column_letter
from typing import List, Union, Dict, Any, Callable, NamedTuple
import json

from EBCommons.paths_and_files import filter_files_with_patterns_and_extensions
from EBCommons.prog_helper import LocalError, get_fun_ref, log_exit, log_init, log_write

__author__ = 'Emmanuel Barillot'

# constantes
# pattern pour les noms de fichiers à rechercher (syntaxe analogue au ls du shell Unix)
PATTERN_INGNEGS_LOG = r"chg*.log"
# encoding des fichiers log à analyser, par defaut
DEFAULT_LOG_ENCODING = r'iso-8859-15'

# Class: Contient les informations d'un compteur
CompteurOne = NamedTuple(b'CompteurOne', [(b'num', int), (b'name', unicode), (b'value', int)])

CompteursFichierBean = NamedTuple(b'CptFi',
                                  [(b'fichier', unicode),
                                   (b'remettant', unicode),
                                   (b'date_run', unicode),
                                   (b'compteurs', Dict)])


# cptfi = CompteursFichierBean(fichier='fic', remettant='INTRUM', date_run='2017/08/01 10:00:00', compteurs=dict())


# TODO pour remplacer CompteursFichiers
# l'idée est d'éviter tout le code inutile
class CompteursFichierExtended(CompteursFichierBean):
    pass


class CompteursFichier(object):
    """
    Contient tous les compteurs associés à un nom de fichier log
    """
    __slots__ = ['_compteurs_line', 'fields_lengths', '_fichier_log', '_fichier_chg', '_remettant', '_date_run',
                 '_target', '_compteurs']

    @property
    def compteurs_line(self):
        """
        :return: la liste des valeurs de tous les champs d'un CompteursFichier, y compris la valeur des compteurs
        """
        return [self.fichier_log,
                self.fichier_chg,
                self.remettant,
                self.date_run] + \
               [val for key, val in self.compteurs_sorted()]

    @property
    def fields_lengths(self):
        """
        :return: la liste des longueurs (en caracteres) des valeurs de tous les champs
        """
        return [len(self.fichier_log),
                len(self.fichier_chg),
                len(self.remettant),
                len(self.date_run)] + \
               [len(str(val)) for key, val in self.compteurs_sorted()]

    @property
    def fichier_log(self):
        # type: () -> unicode
        return os.path.basename(self._fichier_log)

    @property
    def fichier_chg(self):
        # type: () -> unicode
        return self._fichier_chg

    @fichier_chg.setter
    def fichier_chg(self, fichier_chg):
        # type: (unicode) -> None
        self._fichier_chg = fichier_chg

    @property
    def remettant(self):
        # type: () -> unicode
        return self._remettant

    @remettant.setter
    def remettant(self, remettant):
        # type: (unicode) -> None
        self._remettant = remettant

    @property
    def date_run(self):
        # type: () -> unicode
        return self._date_run

    @date_run.setter
    def date_run(self, date_run):
        # type: (unicode) -> None
        self._date_run = date_run

    @property
    def target(self):
        # type: () -> unicode
        return self._target

    @target.setter
    def target(self, target):
        # type: (unicode) -> None
        self._target = target

    def __init__(self, fichier_log, fichier_chg=None, remettant=None, compteurs=None, date_run=None, target=None):
        # type: (unicode, unicode, unicode, Union[List[CompteurOne], None], unicode) -> None
        self._fichier_log = fichier_log
        self._fichier_chg = fichier_chg
        self._remettant = remettant
        self._date_run = date_run
        self._target = target
        self._compteurs = self._compteur_list_to_dict(compteurs)  # type: OrderedDict

    @classmethod
    def _compteur_list_to_dict(cls, compteurs):
        # type: (List[CompteurOne]) -> Dict
        """
        :return: dictionnaire des compteurs du fichier
        """
        if compteurs:
            compteurs_dict = OrderedDict([(compteur.name, compteur.value) for compteur in compteurs])
            return compteurs_dict
        else:
            return OrderedDict()

    def __getitem__(self, item):
        # type: (unicode) -> Any
        """
        :return: dictionnaire des compteurs du fichier
        """
        return self._compteurs[item]

    def compteurs_keys(self):
        # type: () -> List[Any]
        return self._compteurs.keys()

    def compteurs_sorted(self):
        # type: (Callable[Any, Any]) -> List[Any]
        return self._compteurs.items()

    def compteurs(self):
        # type: () -> Dict
        return self._compteurs

    def compteurs_val(self, key):
        # type: (unicode) -> Any
        return self._compteurs[key]

    def add(self, compteur_name=None, compteur_value=None, compteur_one=None):
        # type: (unicode, Any, CompteurOne) -> CompteursFichier
        if compteur_name:
            self._compteurs[compteur_name] = compteur_value
        else:
            self._compteurs[compteur_one.name] = compteur_one.value
        return self

    def print_compteurs(self):
        # type: () -> None
        """affichage des compteurs"""
        log_write("Compteurs du fichier_log {} / fichier chargé {} - {} :".format(self._fichier_log.encode('utf8'),
                                                                                  self._fichier_chg.encode('utf8'),
                                                                                  self._remettant),
                  level=logging.DEBUG)
        for cpt_name, cpt_value in self._compteurs.items():
            log_write("{0:60.60s}: {1:>9}".format(cpt_name, cpt_value), level=logging.DEBUG)

    # TODO transformer la signature pour qu'elle recoive un dict de fonctions associées à un compteur à trouver
    # dans un fichier log
    @classmethod
    def from_file(cls, the_file_name, encoding=DEFAULT_LOG_ENCODING):
        # type: (unicode, unicode) -> CompteursFichier
        """
        Lit les lignes de compteurs d'un fichier log
        et les retourne dans un tuple de tuple de la forme:
            (fichier,(numero,nom,valeur),(...),...)

        :param the_file_name: nom du fichier à analyser
        :param encoding: encodage du fichier à lire
        :return: un tuple des compteurs trouvés
        """

        def parse_date_run(the_line):
            # type: (unicode) -> unicode or None
            re_date_run = re.compile(r"(.*)\|(.*)\|(.*)\|(.*)\|(.*)")  # RE pour trouver la date
            m_date_run = re_date_run.match(the_line)
            if m_date_run is not None:
                _remettant = m_date_run.group(1)
                return _remettant
            return None

        def parse_remettant_conf(this_line):
            # type: (unicode) -> unicode or None
            re_remettant_conf = re.compile(r".*LOG\|supportcodCONF .* : *(.*)")  # RE pour trouver ligne du remettant
            m_remettant_conf = re_remettant_conf.match(this_line)
            if m_remettant_conf is not None:
                _remettant_conf = m_remettant_conf.group(1)
                return _remettant_conf

        def parse_remettant_public(this_line):
            # type: (unicode) -> unicode or None
            re_remettant_public = re.compile(r".*LOG\|supportcod .* : *(.*)")  # RE pour trouver ligne du remettant
            m_remettant_public = re_remettant_public.match(this_line)
            if m_remettant_public is not None:
                _remettant_public = m_remettant_public.group(1)
                return _remettant_public

        def parse_target(this_line):
            # type: (unicode) -> unicode or None
            re_target = re.compile(r".*LOG\|tableCible .* : *(.*)")  # RE pour trouver table coble
            m_target = re_target.match(this_line)
            if m_target is not None:
                _target = m_target.group(1)
                return _target

        def parse_fichier_chg(this_line):
            # type: (unicode) -> unicode or None
            re_target = re.compile(r".*LOG\|nomFichierIn .* : *(.*)")  # RE pour trouver table coble
            m_target = re_target.match(this_line)
            if m_target is not None:
                _target = m_target.group(1).split('/')[-1:]
                return _target[0]

        def parse_nb_compteurs(the_line):
            # type: (unicode) -> int or None
            # re_nb_compteurs = re.compile(".*LOG\|\[(.*)\] Compteurs")
            re_nb_compteurs = re.compile(r".*LOG\|\[(.*)\] \w+$")  # RE pour trouver ligne du nb de compteur
            m_nb_compteurs = re_nb_compteurs.match(the_line)
            if m_nb_compteurs is not None:
                nb_compteurs_match = m_nb_compteurs.group(1)
                return int(nb_compteurs_match)
            return None

        def parse_compteur(the_line):
            # type: (unicode) -> CompteurOne or None
            re_compteur = re.compile(r".*LOG\|\[(.*)\] (.*) \.+ : *(-?\d*)")  # RE pour trouver les lignes des compteurs
            m_compteur = re_compteur.match(the_line)
            if m_compteur is not None:
                g1_compteur = m_compteur.group(1)  # numero du compteur
                g2_compteur = m_compteur.group(2).strip()  # nom du compteur
                g3_compteur = m_compteur.group(3)  # valeur du compteur
                return CompteurOne(int(g1_compteur), g2_compteur if g2_compteur is not None else None, int(g3_compteur))
            return None

        compteurs_fichier = CompteursFichier(fichier_log=os.path.basename(the_file_name))
        with open(the_file_name, b'rb') as f:
            # on lit ligne à ligne et on cherche dans chaque ligne
            #   le motif de la ligne qui contient le nb de compteurs
            #   ou le motif d'une ligne qui contient un compteur
            # recherche de la date dans la 1ere ligne
            date_run = None
            fichier_chg = None
            remettant_conf = None
            remettant_public = None
            nb_compteurs = None
            target = None
            for line in f:
                line_decoded = line.decode(encoding)  # decodage de la ligne vers utf-8 car fichier ouvert en mode 'rb'
                # recherche de la ligne qui contient la date, la 1ere ligne fait l'affaire
                if not date_run:
                    date_run = parse_date_run(line_decoded)
                    if date_run:
                        log_write(the_file_name + ", date_run: " + date_run)
                        compteurs_fichier.date_run = date_run
                # recherche de la ligne qui contient le nom du fichier des données qui ont été chargées
                if not fichier_chg:
                    fichier_chg = parse_fichier_chg(line_decoded)
                    if fichier_chg:
                        log_write(the_file_name + ", fichier_chg: " + fichier_chg)
                        compteurs_fichier.fichier_chg = fichier_chg
                # recherche de la ligne qui contient le remettant
                if remettant_conf is None:
                    remettant_conf = parse_remettant_conf(line_decoded)
                if remettant_public is None:
                    remettant_public = parse_remettant_public(line_decoded)
                if target is None:
                    target = parse_target(line_decoded)
                # la ligne du nb de compteurs
                if nb_compteurs is None:
                    nb_compteurs = parse_nb_compteurs(line_decoded)
                    if nb_compteurs:
                        log_write(the_file_name + ", nb compteurs trouvés: " + str(nb_compteurs))
                # les compteurs
                if nb_compteurs:
                    # si on a trouvé le nb de compteurs, on peut parser les compteurs
                    compteur = parse_compteur(line_decoded)
                    if compteur is not None:
                        compteurs_fichier.add(compteur_one=compteur)
            compteurs_fichier.remettant = remettant_conf if remettant_conf else remettant_public
            compteurs_fichier.target = target if target else None
            if compteurs_fichier.remettant:
                log_write(the_file_name + ", remettant: " + compteurs_fichier.remettant)
        return compteurs_fichier


class CompteursFichierColl(object):
    """
    Collection de CompteursFichier
    """
    __slots__ = ['_coll', '_lmax', '_normalized']

    def __init__(self, compteurs_fichiers):
        # type: (List[CompteursFichier]) -> CompteursFichierColl
        self._coll = compteurs_fichiers     # type: List[CompteursFichier]
        self._lmax = []
        self._normalized = False

    def __getitem__(self, item):
        return self._coll[item]

    @property
    def normalized(self):
        return self._normalized

    @normalized.setter
    def normalized(self, normalized):
        self._normalized = normalized

    def title_line(self, corresp=None):
        # type: (CompteursCorrespondance or None) -> List[unicode]
        if corresp:
            # avec une correspondance, on peut fabriquer une liste ordonnée des noms des champs
            # à partir du numéro qui est attribué à chaque compteur normalisé dans la correspondance
            compteurs_names = corresp.get_compteurs_names_norm_sorted()
        else:
            # sans correspondance, on fabrique une liste ordonnée des noms des champs qu'on trouve dans les
            # compteurs présents dans les fichiers
            compteurs_names = []
            for compteurs_fichier in self._coll:
                compteurs_names += compteurs_fichier.compteurs().keys()
                compteurs_names = list(set(compteurs_names))

        return ['Fichier log', 'Fichier chargé', 'Remettant', 'Date chargement'] + compteurs_names

    @classmethod
    def from_many_files(cls, path_src_file, file_name_list, encoding=DEFAULT_LOG_ENCODING):
        # type: (unicode, List[unicode], unicode) -> CompteursFichierColl
        """
        Lit les lignes de compteurs d'une liste de fichiers log et les retourne dans une liste de CompteursFichier
        :param path_src_file: chemin vers les fichiers à analyser
        :param file_name_list: liste des noms des fichiers à analyser (tous dans le même répertoire path_src)
        :param encoding: encodage à utiliser pour lire les fichiers
        """
        return cls(compteurs_fichiers=map(lambda x: CompteursFichier.from_file(the_file_name=x, encoding=encoding),
                                          [os.path.join(path_src_file, file_name) for file_name in file_name_list]))

    def remettants_names(self):
        # type: () -> List[unicode]
        return sorted(set([cpt.remettant for cpt in self._coll]))

    def upd_lmax(self, field_lengths):
        # type: (List[int]) -> None
        if self._lmax is None or self._lmax == []:
            self._lmax = [1]*len(field_lengths)
        self._lmax = map(max, zip(self._lmax, field_lengths))

    def get_compteurs_fichier_by_remettant(self, remettant):
        # type: (unicode) -> List[CompteursFichier]
        """
        :param remettant: le nom du remettant dont on va retourner les compteurs présents dans la série de fichiers
        qui le concernent
        :return: liste de CompteursFichier triée selon la date_run présente dans chaque CompteursFichier
        """
        non_sorted = filter(lambda x: x.remettant == remettant and x.target == 'INFNEGS', self._coll)

        def format_date(text):
            # date en entrée  DD/MM/YYYY HH:MI:SS
            # date en sortie  YYYY:MM:DD:HH:MI:SS
            text2 = text.split(' ')[0].split('/') + [text.split(' ')[1]]
            return ':'.join([text2[ii] for ii in [2, 1, 0, 3]])

        return sorted(non_sorted, key=lambda x: format_date(x.date_run))

    def add_variation(self):
        """
        Ajoute les valeurs du champ "Variation nb lignes en base" si elles ne sont pas présentes
        """
        for compteurs_fichier in self._coll:
            if not compteurs_fichier.compteurs()["Variation nb lignes en base"]:
                compteurs_fichier.compteurs()["Variation nb lignes en base"] = \
                    compteurs_fichier.compteurs()["Stockés après traitement du fichier"] - \
                    compteurs_fichier.compteurs()["Stockés avant traitement du fichier"]

    def normalize(self):
        # type: () -> CompteursFichierColl
        """
        Normalise les séries de compteurs associés à un ensemble de fichiers
        Tous les fichiers auront artificiellement une valeur (éventuellement 0)
        pour l'ensemble des compteurs présents dans tous les fichiers.
        :return: une collection de CompteursFichier normalisée
        """
        compteur_name_set = set([compteur.name() for compteurs_fichier in self._coll
                                 for compteur in compteurs_fichier.compteurs()])
        compteurs_fichier_list_out = list()
        for compteurs_fichier in self._coll:
            compteurs_normalises_fichier = CompteursFichier(fichier_log=compteurs_fichier.fichier_log,
                                                            fichier_chg=compteurs_fichier.fichier_chg,
                                                            remettant=compteurs_fichier.remettant,
                                                            date_run=compteurs_fichier.date_run,
                                                            target=compteurs_fichier.target,
                                                            compteurs=None)
            for compteur_name in compteur_name_set:
                compteur_value = compteurs_fichier[compteur_name] \
                    if compteur_name in compteurs_fichier.compteurs_keys() else None
                compteurs_normalises_fichier.add(compteur_name=compteur_name, compteur_value=compteur_value)
            compteurs_fichier_list_out += [compteurs_normalises_fichier]
        normalized_coll = CompteursFichierColl(compteurs_fichier_list_out)
        normalized_coll.normalized = True
        normalized_coll.add_variation()
        return normalized_coll

    def normalize_corresp(self, corresp=None):
        # type: (CompteursCorrespondance) -> CompteursFichierColl
        """
        Normalise les séries de compteurs associés à un ensemble de fichiers
        Tous les fichiers auront artificiellement une valeur (éventuellement 0)
        pour l'ensemble des compteurs présents dans tous les fichiers.
        :param corresp: correspondances de compteurs
        :return: une collection de CompteursFichier normalisée
        """
        # si aucune correspondance disponible, on normalise sans
        if corresp is None:
            return self.normalize()

        compteurs_fichier_list_out = list()
        # on parcourt tous les fichiers
        for compteurs_fichier in self._coll:
            compteurs_normalises = CompteursFichier(fichier_log=compteurs_fichier.fichier_log,
                                                    fichier_chg=compteurs_fichier.fichier_chg,
                                                    remettant=compteurs_fichier.remettant,
                                                    date_run=compteurs_fichier.date_run,
                                                    target=compteurs_fichier.target,
                                                    compteurs=None)
            # on parcourt tous les noms de compteurs normalisés
            for compteur_name_norm in corresp.get_compteurs_names_norm_sorted():
                compteur_value = None
                # on parcourt tous les alias de compteurs de fichiers associés au nom de compteur normalisé
                for compteur_name_alias in corresp.get_alias(compteur_name_norm):
                    # si un alias existe dans les compteurs du fichier en cours, alors on garde ses valeurs
                    re_compiled = re.compile(compteur_name_alias)
                    for compteur_name_fichier in compteurs_fichier.compteurs_keys():
                        if re_compiled.match(compteur_name_fichier) is not None:
                            # le nom du compteur dans le fichier doit contenir l'alias
                            # l'égalité stricte des noms n'est pas exigée
                            # on cumule les valeurs des compteurs de tous les alias
                            if not compteur_value:
                                compteur_value = 0
                            compteur_value += compteurs_fichier[compteur_name_fichier]
                compteurs_normalises.add(compteur_name=compteur_name_norm, compteur_value=compteur_value)
            self.upd_lmax(compteurs_normalises.fields_lengths)
            compteurs_fichier_list_out += [compteurs_normalises]
        normalized_coll = CompteursFichierColl(compteurs_fichier_list_out)
        normalized_coll.normalized = True
        normalized_coll.add_variation()
        return normalized_coll

    def to_excel_file(self, excel_file_full_name, p_corresp_file_name=None):
        # type: (unicode, unicode or None) -> None
        """
        Crée un fichier Excel à partir des compteurs d'un ou de plusieurs fichiers.
        :param excel_file_full_name: nom du fichier Excel en sortie, chemin compris
        :param p_corresp_file_name: nom du fichier qui contient les correspondances
        :return: None
        """

        # normalisation des compteurs
        corresp = CompteursCorrespondance.init_from_file(p_corresp_file_name)
        compteurs_coll = self.normalize_corresp(corresp)

        def add_title_style(wb):
            # creation d'un style personnel nommé, de façon à l'affecter à une ligne ou une colonne entière
            title_style = NamedStyle(name="title_style")
            title_style.font = Font(bold=True, size=14)
            bd = Side(style='thin', color="000000")
            title_style.border = Border(left=None, top=bd, right=None, bottom=bd)
            title_style.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
            wb.add_named_style(title_style)

        def format_feuille(p_feuille):
            # ajustement éventuel de la largeur de chaque colonne
            for i in range(1,p_feuille.max_column+1):
                p_feuille.column_dimensions[get_column_letter(i)].width = max(int(self._lmax[i-1]*1.1),10)

            # p_feuille.column_dimensions[get_column_letter(1)].width = 54
            # p_feuille.column_dimensions[get_column_letter(2)].width = 54
            # for i in range(3, p_feuille.max_column):
            #     p_feuille.column_dimensions[get_column_letter(i + 1)].width = 13

            # figer les volets sur la première ligne
            p_feuille.freeze_panes = p_feuille.cell(row=2, column=1)

            # application du style au titre
            for cells in p_feuille.iter_rows(min_row=1, max_row=1, min_col=1, max_col=p_feuille.max_column):
                for cell in cells:
                    cell.style = 'title_style'

        # création du document Excel en mémoire et des feuilles: une feuille par remettant
        workbook = ExcelWorkbook()
        add_title_style(workbook)
        feuille = workbook.active  # grab the active worksheet, tjrs présente par défaut

        # creation d'une feuille par remettant
        first_remettant = 1
        for remettant in compteurs_coll.remettants_names():
            if first_remettant:
                first_remettant = None
                # 1ere feuille associée au 1er remettant de la liste
                feuille.title = remettant
            else:
                # création d'une nouvelle feuille
                feuille = workbook.create_sheet(remettant)

            # ajout des en-têtes, à partir des libellés
            # Les compteurs sont triés en fonction du numéro qui leur est attribué dans la correspondance
            feuille.append(compteurs_coll.title_line(corresp))
            # ajout des valeurs dans les lignes suivantes pour les lignes qui concernent un remettant donné
            for compteurs_fichier in compteurs_coll.get_compteurs_fichier_by_remettant(remettant):
                next_line = compteurs_fichier.compteurs_line
                feuille.append(next_line)
            format_feuille(feuille)

        # création matérielle du fichier résultant
        try:
            workbook.save(excel_file_full_name)
        except Exception as e:
            log_write(LocalError(e).__str__(), level=logging.ERROR)


class CompteursCorrespondance(object):
    """
    Contient une correspondance entre noms de compteurs. Une correspondance est un dictionnaire qui fait correspondre
    un nom normalisé avec un ensemble de noms à chercher dans les fichiers de log.
    On associe un nom de compteur normalisé qui apparaitra dans le fichier de synthèse. Ca permet de générer une
    synthèse des compteurs tout en gérant l'évolution des noms de compteurs au fil des évolutions des programmes
    de chargement.
    """
    __slots__ = ['_corresp', '_json_file']
    COMPTEURS_TAG = 'compteurs'
    NUM_TAG = 'num'
    ALIAS_TAG = 'alias'

    def __init__(self, json_obj, json_file=None):
        # type: (Dict[Any]) -> CompteursCorrespondance
        self._corresp = json_obj  # type: Dict[Any]
        self._json_file = json_file  # type: unicode

    def get_alias(self, compteur_name):
        # type: (unicode) -> List[unicode]
        return self._corresp[self.COMPTEURS_TAG][compteur_name][self.ALIAS_TAG]

    def get_compteurs_names_norm(self):
        # type: (unicode) -> List[unicode]
        return self._corresp[self.COMPTEURS_TAG].keys()

    def get_compteurs_names_norm_sorted(self):
        # type: (unicode) -> List[unicode]
        """
        :return: la liste des noms des compteurs normalisés triée selon le numéro attribué dans la correspondance
        """
        return [key for key, val in sorted(self._corresp[self.COMPTEURS_TAG].items(), key=lambda t: t[1][self.NUM_TAG])]

    def get_fun_cmp(self):
        # type: () -> Callable[Any, Any]
        """
        :return: une fonction de comparaison entre deux noms de compteurs, qui permet de les ordonner en fonction
        du numéro de compteur qui est attribué à chacun dans la correspondance
        """

        def _cmp(compteur_name_1, compteur_name_2):
            num_1 = self._corresp[self.COMPTEURS_TAG][compteur_name_1][self.NUM_TAG]
            num_2 = self._corresp[self.COMPTEURS_TAG][compteur_name_2][self.NUM_TAG]
            if num_1 > num_2:
                return 1
            elif num_1 < num_2:
                return -1
            else:
                return 0

        return _cmp

    @classmethod
    def init_from_file(cls, file_name):
        # type: (unicode) -> CompteursCorrespondance
        """
        Construit une instance d'une correspondance de compteurs à partir d'un fichier de paramétrage
        :param file_name: nom du fichier qui contient la correspondance entre les noms de compteurs
        :return: la correspondance, en fait un dictionnaire de plusieurs correspondances vers un seul ensemble
        de compteurs.
        """
        with open(file_name) as f:
            json_obj = json.load(f)
        cls._valid_json(json_obj, json_file=file_name)
        return cls(json_obj=json_obj, json_file=file_name)  # appelle __init__()

    @classmethod
    def _valid_json(cls, corresp_json, json_file):
        # type: (Dict, unicode) -> None
        """
        Valide le json qui contient les correspondances:
         - un numéro unique doit être associé à chaque entrée
         - un numéro de version doit être présent
        :param: corresp_json
        :param: json_file
        :return: None
        """
        if "Version" not in corresp_json.keys():
            raise Exception("Aucune indication de version dans le fichier {]".format(json_file))
        # controle du nb de numéros
        nums = [corresp_json[cls.COMPTEURS_TAG][key][cls.NUM_TAG] for key in corresp_json[cls.COMPTEURS_TAG].keys()
                if isinstance(corresp_json[cls.COMPTEURS_TAG][key], dict)
                and cls.NUM_TAG in corresp_json[cls.COMPTEURS_TAG][key]]
        if len(set(nums)) != len(nums):
            raise Exception("Problème dans les numéros d'ordre dans le fichier {]".format(json_file))
        # controle de la présence du champ alias
        for key in corresp_json[cls.COMPTEURS_TAG].keys():
            try:
                corresp_json[cls.COMPTEURS_TAG][key][cls.ALIAS_TAG]
            except LookupError:
                raise Exception("Il manque un champ alias pour la clé {}".format(key))


def call_read_prg_log(dir_name, file_name):
    # type: (unicode, unicode) -> bool
    """
    Appel unitaire de la fonction de recherche des compteurs from_file()

    :param dir_name: répertoire du fichier à analyser
    :param file_name: nom du fichier log à analyser
    :return: True si exécution OK
    """
    the_file_name = os.path.join(dir_name, file_name)
    try:
        cpts = CompteursFichier.from_file(the_file_name)
        cpts.print_compteurs()
        return True
    except Exception as e:
        log_write(LocalError(e).__str__(), level=logging.ERROR)
        return False


def call_read_prg_log_many(dir_name, re_file_name):
    # type: (unicode) -> bool
    """
    Appel unitaire de la fonction de recherche des compteurs from_many_files()

    :param dir_name: répertoire à scruter pour trouver les fichiers log à analyser
    :param re_file_name: expression régulière pour les noms de fichiers à analyser
    :return: bool
    """
    try:
        file_name_list = \
            filter_files_with_patterns_and_extensions(os.listdir(dir_name), include_patterns=(re_file_name,))
        cpt_list = CompteursFichierColl.from_many_files(dir_name, file_name_list, DEFAULT_LOG_ENCODING)
        for cpt in cpt_list:
            cpt.print_compteurs()
        return True
    except Exception as e:
        log_write(LocalError(e).__str__(), level=logging.ERROR)
        return False


def call_excel_write_log_cpt(dir_name, excel_file_name):
    # type: (unicode, unicode) -> bool
    try:
        file_name = os.path.join(dir_name, excel_file_name)
        compteurs_coll = CompteursFichierColl(
            [
                CompteursFichier(fichier_log='fichier_1'
                                 , remettant='Remettant 1'
                                 , compteurs=map(lambda x: CompteurOne(*x)
                                                 , [
                                                     (' 0', 'cpt1', 12)
                                                     , (' 1', 'cpt2', 11)
                                                     , (' 2', 'cpt3', 9)
                                                     , (' 3', 'cpt4', 8)
                                                     ]
                                                 )
                                 )
                , CompteursFichier(fichier_log='fichier_2'
                                   , remettant='Remettant 2'
                                   , compteurs=map(lambda x: CompteurOne(*x)
                                                   , [
                                                       (' 0', 'cpt1', 13)
                                                       , (' 1', 'cpt2', 15)
                                                       , (' 2', 'cpt3', 6)
                                                       , (' 3', 'cpt4', 7)
                                                       ]
                                                   )
                                   )
                ]
            )

        compteurs_coll.to_excel_file(file_name, p_corresp_file_name=None)
        return True
    except Exception as e:
        log_write(LocalError(e).__str__(), level=logging.ERROR)
        return False


def call_read_prg_log_to_excel(dir_name, file_name, excel_file_name):
    # type: (unicode, unicode, unicode) -> bool
    """
    Lit un fichier de logs et ecrit les compteurs extraits dans un fichier Excel
    :param dir_name: répertoire où se trouve le fichier
    :param file_name: le nom du fichier
    :param excel_file_name: le nom du fichier Excel
    :return: bool
    """
    the_file_name = os.path.join(dir_name, file_name)
    the_excel_file_name = os.path.join(dir_name, excel_file_name)
    try:
        cpts = CompteursFichier.from_file(the_file_name)
        cpts.print_compteurs()
        CompteursFichierColl([cpts]).to_excel_file(the_excel_file_name)
        return True
    except Exception as e:
        log_write(LocalError(e).__str__(), level=logging.ERROR)
        return False


def call_read_prg_log_to_excel_many(p_path_src, file_name_re, path_dest, excel_file_name,
                                    p_corresp_file_name, encoding_src=DEFAULT_LOG_ENCODING):
    # type: (unicode, unicode, unicode, unicode, unicode) -> bool
    """
    Lit une liste de fichiers de logs et ecrit les compteurs extraits dans un fichier Excel
    :param p_path_src: répertoire où se trouvent les fichiers
    :param file_name_re: le motif des noms de fichiers en expression régulière nom du fichier
    :param encoding_src: encodage des fichiers logs (par ex iso-8859-15 sous Linux)
    :param path_dest: répertoire de destination du fichier Excel
    :param excel_file_name: le nom du fichier Excel produit
    :param p_corresp_file_name: nom du fichier des correspondances de compteurs
    :return: True si succès
    """
    file_name_re = PATTERN_INGNEGS_LOG if file_name_re is None else file_name_re
    # on recupere les noms de fichiers à traiter
    file_name_list = filter_files_with_patterns_and_extensions(os.listdir(p_path_src), include_patterns=(file_name_re,))
    try:
        cpt_list = CompteursFichierColl.from_many_files(p_path_src, file_name_list, encoding_src)
        # affichage des compteurs
        if cpt_list:
            for cpt in cpt_list:
                cpt.print_compteurs()
            cpt_list.to_excel_file(os.path.join(path_dest, excel_file_name), p_corresp_file_name)
        return True
    except Exception as e:
        log_write(LocalError(e).__str__(), level=logging.ERROR)
        raise
        # return False


#####################
# programme principal
#####################
def launch(p_path_src, log_file_name, encoding_src, path_dest, excel_file_name, p_corresp_file_name):
    # type: (unicode, unicode, unicode, unicode, unicode) -> List(bool)
    """
    Lance les fonctions pour test / validation

    :param p_path_src: repertoire source où se trouvent les fichiers log à analyser
    :param log_file_name: nom du fichier à analysez quand execution sur un seul fichier
    :param encoding_src: encodage des fichiers log
    :param path_dest: répertoire de destination pour le fichier Excel
    :param excel_file_name: nom du fichier Excel
    :param p_corresp_file_name: nom du fichier des correspondances de compteurs
    :return: liste de bool
    """
    # liste de fonctions à tester
    dict_of_funs_to_test = {
        r'call_read_prg_log'                : (p_path_src, log_file_name)
        , r'call_excel_write_log_cpt'       : (p_path_src, excel_file_name)
        , r'call_read_prg_log_to_excel'     : (p_path_src, log_file_name, path_dest, excel_file_name)
        , r'call_read_prg_log_many'         : (p_path_src, PATTERN_INGNEGS_LOG)
        , r'call_read_prg_log_to_excel_many': (
            p_path_src, PATTERN_INGNEGS_LOG, path_dest, excel_file_name, p_corresp_file_name, encoding_src)
        }

    # liste des références des fonctions
    funs_available = map(lambda x: get_fun_ref(x, __name__), dict_of_funs_to_test.keys())
    # Autre technique qui utilise eval, peut être dangereuse si le nom de la fonction provient d'une entrée utilisateur
    # funs_available = map(eval, dict_of_funs_to_test.keys())

    # la liste des arguments des fonctions testées
    for fun in funs_available:
        if "inspect" not in globals().keys():
            from inspect import getargspec
            args, varargs, varkw, defaults = getargspec(fun)
            log_write("{}:".format(fun))
            log_write("{:10s}: {}".format("args", args))
            log_write("{:10s}: {}".format("varargs", varargs))
            log_write("{:10s}: {}".format("varkw", varkw))
            log_write("{:10s}: {}".format("defaults", defaults))

    # funs_to_test = [call_read_prg_log] # la fonction à tester
    # funs_to_test = [call_excel_write_log_cpt] # la fonction à tester
    # funs_to_test = [call_read_prg_log_to_excel] # la fonction à tester
    # funs_to_test = [call_read_prg_log_many]
    funs_to_test = [call_read_prg_log_to_excel_many]

    # lancement des tests
    call_return = []
    for fun_to_test in funs_to_test:
        log_write(">>>>> testing: {}".format(fun_to_test.func_name))
        # l'étoile devant *dict_of_funs_to_test[fun_to_test.func_name]
        # sert à "unpacker" le tuple stocké dans chaque entrée de ce dictionnaire
        # il faut unpacker le tuple avant d'en passer les éléments à la fonction appelée
        call_return.append((fun_to_test.func_name,
                            "OK" if fun_to_test(*dict_of_funs_to_test[fun_to_test.func_name]) else "KO"))
    return call_return


##################################
# lancement du programme principal
##################################
if __name__ == "__main__":
    log_init(level=logging.DEBUG)
    log_write(">>>>>>>>>>>>>>>>>>>> Nouvelle série de tests <<<<<<<<<<<<<<<<<<<<")

    # path_root = r"D:\Documents\Projets\work\Infnegs_logs\2017-08\qqn"
    # path_root = r"D:\Documents\Projets\work\Infnegs_logs\2017-08"
    # path_root = r"D:\Documents\Projets\work\Infnegs_logs\2017-09"
    path_root = r"D:\Documents\Projets\work\Infnegs_logs\2017-10"
    path_src = path_root
    path_dest_for_excel = path_root
    one_log_file_name = "chgInfnegs_201610031669949.log"
    excel_out_filename = "compteurs.xlsx"
    corresp_file_name = r'compteurs_correspondances_V3.json'

    res = launch(path_src, one_log_file_name, DEFAULT_LOG_ENCODING, path_dest_for_excel, excel_out_filename,
                 corresp_file_name)
    if res is not None:
        for elem in res:
            log_write("{}() => {}".format(unicode(elem[0]), unicode(elem[1])))

    log_exit()
